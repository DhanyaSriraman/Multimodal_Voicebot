# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


from flask import Flask, render_template, jsonify
from flask import request
import requests
import html5lib
import openpyxl
import numpy as np
import pandas as pd
from requests_html import HTMLSession
from bs4 import BeautifulSoup
from pipelines import pipeline
from transformers import pipeline as p
import torch
# from googlesearch import search

# app = Flask(__name__)


# @app.route('/')
# def hello_world():
#     return render_template('index.html')

# @app.route('/check',methods=['POST'])
def predict():
    ws = openpyxl.load_workbook('questions_and_answers.xlsx')
    qA = ws['Sheet1']
    session = HTMLSession()
    ws2 = openpyxl.load_workbook('articleLinks.xlsx')
    sheet=ws2.active

    last_column = sheet.max_column

# Get the values in the last column, skipping the first row
    values = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
       for cell in row:
         values.append(cell.value)

    for url in values :
    
      with session.get(url) as r:

        selector = 'p'
        paragraph = r.html.find(selector, first=False)

        text = " ".join([ p.text for p in paragraph])
        text1 = "sam oru loosu"
    # try:

    # except ImportError:
    #     print("No module named 'google' found")

#     query = text.split("\n")[0]

#     link_list = []
#     print(str2)
#     for j in search(str2, tld="co.in", num=10, stop=10, pause=1):
#         link_list.append(j)
#     print(link_list)

        nlp = pipeline("multitask-qa-qg")

#     text = """Several social media users are circulating a photo, which we have received on our Whatsapp tipline, too, claiming a “Congress conspiracy” behind the recent controversial BBC documentary on Prime Minister Narendra Modi, while stating that Congress leader Rahul Gandhi met the documentary’s producer during his time in the UK last year.
# The archived versions of the tweets can be seen here, here and here.
# Newschecker saw that the people in the photo with Rahul Gandhi were Indian entrepreneur Sam Pitroda, along with UK MP and former Labour leader Jeremy Corbyn. Upon doing a relevant keyword search, we were led to multiple media reports, seen here, here and here, on the 2022 meeting in London, which had sparked a political row.
# According to an India Today report, dated May 24, 2022, the BJP criticised Gandhi over his photo with Corbyn, following the party’s rift with the “anti-India” MP stemming from his “numerous tweets on the situation in Jammu and Kashmir, which India has persistently maintained is an internal matter”. Pitroda, a close aide of the Gandhi family for years, reportedly told India Today TV: “He (Corbyn) is a personal friend of mine and came over for a cup of tea at the hotel. There’s nothing political about this.”
# We learnt that the photo was first tweeted out by the Indian Overseas Congress on May 23, 2022.
# Our chairman @sampitroda with @RahulGandhi in London 🥰 pic.twitter.com/veyWjx1bpL
#  We then looked up the credits of “India: The Modi Question”, the first episode of which was aired on January 17, 2023, on IMDb and BBC. We learnt that the series producer was Richard Cookson and executive producer Mike Radford. Taking a cue from this, we ran a keyword search for “Rahul Gandhi Richard Cookson Mike Radford”, which did not throw up any relevant results or photos of such a meeting.
#
# We reached out to the BBC, where a spokesperson said, “Nobody from the production team met with Rahul Gandhi.”"""
        ans = nlp(text)

#         print(ans)
        only_question = []
        ques_ans_dict = {}
   
        for dictionary in ans:
          only_question.append(dictionary['question'])
          ques_ans_dict[dictionary['question']] = dictionary['answer']
#         print(ques_ans_dict)
          nextRow=qA.max_row+1
          for key, value in ques_ans_dict.items():
               qA.cell(row=nextRow,column=1,value=key)
               qA.cell(row=nextRow,column=2,value=value)
               nextRow=nextRow+1
          ws.save('questions_and_answers.xlsx')
#     with torch.no_grad():

#         question_answerer = p("question-answering", model='distilbert-base-cased-distilled-squad')

#         context = r"""Several social media users are circulating a photo, which we have received on our Whatsapp tipline, too, claiming a “Congress conspiracy” behind the recent controversial BBC documentary on Prime Minister Narendra Modi, while stating that Congress leader Rahul Gandhi met the documentary’s producer during his time in the UK last year.
# The archived versions of the tweets can be seen here, here and here.
# Newschecker saw that the people in the photo with Rahul Gandhi were Indian entrepreneur Sam Pitroda, along with UK MP and former Labour leader Jeremy Corbyn. Upon doing a relevant keyword search, we were led to multiple media reports, seen here, here and here, on the 2022 meeting in London, which had sparked a political row.
# According to an India Today report, dated May 24, 2022, the BJP criticised Gandhi over his photo with Corbyn, following the party’s rift with the “anti-India” MP stemming from his “numerous tweets on the situation in Jammu and Kashmir, which India has persistently maintained is an internal matter”. Pitroda, a close aide of the Gandhi family for years, reportedly told India Today TV: “He (Corbyn) is a personal friend of mine and came over for a cup of tea at the hotel. There’s nothing political about this.”
# We learnt that the photo was first tweeted out by the Indian Overseas Congress on May 23, 2022.
# Our chairman @sampitroda with @RahulGandhi in London 🥰 pic.twitter.com/veyWjx1bpL
#  We then looked up the credits of “India: The Modi Question”, the first episode of which was aired on January 17, 2023, on IMDb and BBC. We learnt that the series producer was Richard Cookson and executive producer Mike Radford. Taking a cue from this, we ran a keyword search for “Rahul Gandhi Richard Cookson Mike Radford”, which did not throw up any relevant results or photos of such a meeting.

# We reached out to the BBC, where a spokesperson said, “Nobody from the production team met with Rahul Gandhi.
#           """

#         context = str1

#         result = question_answerer(question="Who was Richard Cookson's executive producer?",
#                                    context=context)
#         print(
#             f"Answer: '{result['answer']}', score: {round(result['score'], 4)}, start: {result['start']}, end: {result['end']}")

#     return render_template('index.html', prediction_text='This Article is False :{}'.format(str))


# def check():
#     text = [x for x in request.form.values()]
#     link_list = take_text(text)
#     nlp = pipeline("multitask-qa-qg")
#     question_ans = nlp(text)
#     only_question = []
#     ques_ans_dict = {}
#     ans_all = []
#     for dictionary in question_ans:
#         only_question.append(dictionary['question'])
#         ques_ans_dict[dictionary['question']] = dictionary['answer']
#     query_search_text(link_list, only_question, ans_all)
#     # evaluation wala Karn hai ab
#     print(ques_ans_dict)
#     print(only_question)
#     print(ans_all)
#     return render_template('index.html', prediction_text ='This Article is False :{}'.format("Ho Gaya"))



# def take_text(text):

#     try:
#         from googlesearch import search
#     except ImportError:
#         print("No module named 'google' found")
#     query = text.split("\n")[1]
#     link_list = []
#     for j in search(query, tld="co.in", num=10, stop=10, pause=2):
#         link_list.append(j)
#     return link_list


# def query_search_text(link_list, only_question, ans_all):

#     new_list = link_list[0: 5]
#     for link in new_list:
#         URL = link
#         r = requests.get(URL)
#         soup = BeautifulSoup(r.content, 'html5lib')
#         tags = soup.find_all('h1')
#         #
#         heading = ""
#         i = 0
#         for t in tags:
#             try:
#                 if i == 0:
#                     heading = t.text
#                     i += 1
#                 print(t.text)
#             except:
#                 continue

#         tags1 = soup.find_all('p')
#         para = ""
#         i = 0
#         for t in tags1:
#             try:
#                 print(t.text)
#                 para += t.text
#             except:
#                 continue
#         answer_from_link(para, only_question, ans_all)




# # def generate_questions_ans(text):
# #     nlp = pipeline("multitask-qa-qg")
# #     question_ans = nlp(text)
# #     only_question = []
# #     ques_ans_dict = {}
# #     for dictionary in question_ans:
# #         only_question.append(dictionary['question'])
# #         ques_ans_dict[dictionary['question']] = dictionary['answer']

# def answer_from_link(para, only_question,ans_obtained_all):
#     with torch.no_grad():
#         question_answerer = p("question-answering", model='distilbert-base-cased-distilled-squad')
#         context = para
#         ans = []
#         for question in only_question:
#             result = question_answerer(question = question,context=context)
#             ans.append(result)
#         ans_obtained_all.append(ans)

if __name__ == '__main__':
     predict()